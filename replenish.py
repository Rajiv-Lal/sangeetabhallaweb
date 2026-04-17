import pandas as pd
import subprocess
import sys
import json
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MASTER     = "/Users/rajivlal/Desktop/whatsapp-sender/whatsapp_final.json"
REVIEW     = "/Users/rajivlal/Desktop/whatsapp-sender/daily_review.xlsx"
BATCH_SIZE = 50

def fix_name(val):
    name = str(val or '').strip()
    if name in ('', 'nan', 'None'):
        return ''
    for title in ['Dr.', 'Dr ', 'Prof.', 'Prof ', 'Mr.', 'Mr ', 'Mrs.', 'Mrs ', 'Ms.', 'Ms ']:
        if name.lower().startswith(title.lower()):
            name = name[len(title):].strip()
    return name.split()[0] if name else ''

def wa_colour(phone):
    p = str(phone)
    if p.startswith('91') and len(p) == 12:
        return 'E8F5E9'
    if p.startswith('1') and len(p) == 11:
        return 'E3F2FD'
    return 'FFF9C4'

thin   = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_MAIN   = "1F4E79"
HDR_ACTION = "375623"

out_cols = ['first_name', 'Name', 'Phone Number', 'Email', 'Source', 'review_note']

col_widths = {
    'first_name': 16, 'Name': 24, 'Phone Number': 18,
    'Email': 28, 'Source': 20, 'review_note': 30
}
action_cols = ['first_name', 'review_note']

# Read current daily_review — skip banner (row 1) and use row 2 as header
try:
    review = pd.read_excel(REVIEW, dtype=str, header=1)
    # Rename columns if they came back as unnamed
    review.columns = [str(c).strip() for c in review.columns]
except Exception as e:
    print(f"Could not read daily_review.xlsx: {e}")
    sys.exit(1)

# Count valid rows (have a phone number)
valid_rows = review[review['Phone Number'].notna() & (review['Phone Number'].str.strip() != '') & (review['Phone Number'].str.strip() != 'nan')]
current_count = len(valid_rows)
print(f"Current daily_review has {current_count} valid contacts")

if current_count >= BATCH_SIZE:
    print(f"Already at {BATCH_SIZE}. No replenishment needed.")
    sys.exit(0)

needed = BATCH_SIZE - current_count
print(f"Need {needed} more contacts to reach {BATCH_SIZE}")

# Read master list from JSON
with open(MASTER) as f:
    master_records = json.load(f)
master = pd.DataFrame(master_records)
if 'status' not in master.columns:
    master['status'] = 'pending'
master['status'] = master['status'].fillna('pending')
master.loc[master['status'].str.strip() == '', 'status'] = 'pending'

# Phones already in review
review_phones = set(valid_rows['Phone Number'].str.strip().str[-10:].tolist())

# Get pending contacts not already in review
pending = master[
    (master['status'].str.strip() == 'pending') &
    (~master['Phone Number'].str.strip().str[-10:].isin(review_phones))
].copy()

if len(pending) == 0:
    print(f"No more pending contacts. Proceeding with {current_count} contacts.")
    sys.exit(0)

top_up = pending.head(needed).copy()
top_up['first_name']   = top_up['Name'].apply(fix_name)
top_up['review_note']  = ''

print(f"Adding {len(top_up)} contacts")

# Align top_up to out_cols
for col in out_cols:
    if col not in top_up.columns:
        top_up[col] = ''

# Align valid_rows to out_cols
for col in out_cols:
    if col not in valid_rows.columns:
        valid_rows = valid_rows.copy()
        valid_rows[col] = ''

# Combine existing + new
combined = pd.concat([valid_rows[out_cols], top_up[out_cols]], ignore_index=True)
total    = len(combined)
print(f"Total contacts in daily_review: {total}")

# Write new daily_review.xlsx
wb = Workbook()
ws = wb.active
ws.title = f"Review {date.today().strftime('%d-%b')}"
ws.freeze_panes = "A3"

# Row 1 — instruction banner
ws.merge_cells("A1:I1")
inst = ws["A1"]
inst.value = (
    f"TODAY: {date.today().strftime('%d %b %Y')}  |  "
    f"Set variant (A/B/C), sign_off and laptop for each row.  |  "
    f"Edit first_name if wrong.  |  "
    f"Total: {total} contacts ready to send."
)
inst.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
inst.fill      = PatternFill("solid", start_color="375623")
inst.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 22

# Row 2 — headers
for col_idx, col in enumerate(out_cols, 1):
    c = ws.cell(row=2, column=col_idx, value=col)
    hdr = HDR_ACTION if col in action_cols else HDR_MAIN
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", start_color=hdr)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col, 14)
ws.row_dimensions[2].height = 28

# Data rows from row 3
for row_idx, (_, row) in enumerate(combined.iterrows(), 3):
    phone = str(row.get('Phone Number', '') or '').strip()
    bg    = wa_colour(phone)
    for col_idx, col in enumerate(out_cols, 1):
        val = row.get(col, '')
        if pd.isna(val) or str(val) == 'nan':
            val = ''
        else:
            val = str(val).strip()
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font          = Font(name="Arial", size=10)
        c.alignment     = Alignment(horizontal="left", vertical="center")
        c.border        = border
        c.fill          = PatternFill("solid", start_color=bg)
        if col == 'Phone Number' and val:
            c.value     = str(val)
            c.data_type = 's'

# No dropdowns needed — single template, fixed sign-off

wb.save(REVIEW)
print(f"Saved daily_review.xlsx with {total} contacts.")

subprocess.Popen(['open', REVIEW])
